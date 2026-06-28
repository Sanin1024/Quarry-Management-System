import io
import os
from datetime import datetime, date, timedelta
from flask import Blueprint, request, jsonify, Response, send_file
import pandas as pd
from models import db, WorkingPeriod
from routes.periods import serialize_period

reports_bp = Blueprint('reports', __name__)

def get_filtered_query(quarry, report_type, date_from_str, date_to_str):
    query = WorkingPeriod.query
    
    # 1. Quarry Filter
    if quarry and quarry != 'both':
        query = query.filter(WorkingPeriod.quarry == quarry)
        
    # 2. Date Filter Range
    # Find "today" reference based on latest record in database to support seeded history
    latest_record = WorkingPeriod.query.order_by(WorkingPeriod.day_from.desc()).first()
    ref_today = latest_record.day_to if latest_record else date.today()
    
    start_date = None
    end_date = None
    
    if report_type == 'daily':
        # Find records covering ref_today
        query = query.filter(WorkingPeriod.day_from <= ref_today, WorkingPeriod.day_to >= ref_today)
        start_date = ref_today
        end_date = ref_today
    else:
        if report_type == 'weekly':
            start_date = ref_today - timedelta(days=7)
            end_date = ref_today
        elif report_type == 'monthly':
            start_date = ref_today - timedelta(days=30)
            end_date = ref_today
        elif report_type == 'yearly':
            start_date = ref_today - timedelta(days=365)
            end_date = ref_today
        elif report_type == 'custom':
            if date_from_str:
                try:
                    start_date = datetime.strptime(date_from_str, '%Y-%m-%d').date()
                except ValueError:
                    pass
            if date_to_str:
                try:
                    end_date = datetime.strptime(date_to_str, '%Y-%m-%d').date()
                except ValueError:
                    pass
                    
        if start_date:
            query = query.filter(WorkingPeriod.day_from >= start_date)
        if end_date:
            query = query.filter(WorkingPeriod.day_to <= end_date)
            
    return query, start_date or ref_today, end_date or ref_today

# GET /api/reports/generate
@reports_bp.route('/generate', methods=['GET'])
def generate_report():
    report_type = request.args.get('type', 'custom')
    quarry = request.args.get('quarry', 'both')
    date_from_str = request.args.get('date_from')
    date_to_str = request.args.get('date_to')

    query, start_date, end_date = get_filtered_query(quarry, report_type, date_from_str, date_to_str)
    periods = query.order_by(WorkingPeriod.day_from.asc()).all()

    total_working_days = sum(p.working_days for p in periods)
    total_expense = sum(p.total_expense for p in periods)
    total_revenue = sum(p.total_revenue for p in periods)
    total_land_lease = sum(p.land_lease_value for p in periods)
    net_value = sum(p.net_value for p in periods)
    total_received = sum(p.received_amount for p in periods)
    total_outstanding = sum(p.balance_outstanding for p in periods)

    return jsonify({
        'summary': {
            'total_working_days': total_working_days,
            'total_expense': total_expense,
            'total_revenue': total_revenue,
            'total_land_lease': total_land_lease,
            'net_value': net_value,
            'total_received': total_received,
            'total_outstanding': total_outstanding
        },
        'items': [serialize_period(p) for p in periods],
        'filters': {
            'type': report_type,
            'quarry': quarry,
            'date_from': start_date.isoformat(),
            'date_to': end_date.isoformat()
        }
    })

# GET /api/reports/export
@reports_bp.route('/export', methods=['GET'])
def export_report():
    export_format = request.args.get('format', 'excel')
    report_type = request.args.get('type', 'custom')
    quarry = request.args.get('quarry', 'both')
    date_from_str = request.args.get('date_from')
    date_to_str = request.args.get('date_to')

    query, start_date, end_date = get_filtered_query(quarry, report_type, date_from_str, date_to_str)
    periods = query.order_by(WorkingPeriod.day_from.asc()).all()

    # Aggregate summaries
    total_working_days = sum(p.working_days for p in periods)
    total_expense = sum(p.total_expense for p in periods)
    total_revenue = sum(p.total_revenue for p in periods)
    total_land_lease = sum(p.land_lease_value for p in periods)
    net_value = sum(p.net_value for p in periods)
    total_received = sum(p.received_amount for p in periods)
    total_outstanding = sum(p.balance_outstanding for p in periods)

    # Filename prefix
    quarry_slug = 'quarry1' if quarry == 'Quarry 1' else ('quarry2' if quarry == 'Quarry 2' else 'both')
    filename = f"{quarry_slug}_report_{start_date.isoformat()}_to_{end_date.isoformat()}"

    # Build rows
    rows = []
    for p in periods:
        rows.append({
            'ID': p.id,
            'Quarry': p.quarry,
            'Day From': p.day_from.isoformat() if p.day_from else '',
            'Day To': p.day_to.isoformat() if p.day_to else '',
            'Working Days': p.working_days,
            'Labourers': p.num_labourers,
            'Labour Pay': p.labour_pay,
            'Diesel Expense': p.diesel_expense,
            'Spare Parts': p.spare_parts,
            'Fitting Charge': p.fitting_charge,
            'JCB Charge': p.jcb_charge,
            'Cutting Wheel': p.cutting_wheel,
            'Mess Expense': p.mess_expense,
            'Other Expense': p.other_expense,
            'Total Expense': p.total_expense,
            'First Quality Bricks': p.first_quality_bricks,
            'Second Quality Bricks': p.second_quality_bricks,
            'Broken Bricks Loads': p.broken_bricks_loads,
            'Total Revenue': p.total_revenue,
            'Land Lease Value': p.land_lease_value,
            'Net Value': p.net_value,
            'Received Amount': p.received_amount,
            'Balance Outstanding': p.balance_outstanding,
            'Source File': p.source_file or ''
        })

    if rows:
        # Add summary row at the bottom
        totals_row = {
            'ID': 'Total',
            'Quarry': '',
            'Day From': '',
            'Day To': '',
            'Working Days': total_working_days,
            'Labourers': sum(r['Labourers'] for r in rows),
            'Labour Pay': sum(r['Labour Pay'] for r in rows),
            'Diesel Expense': sum(r['Diesel Expense'] for r in rows),
            'Spare Parts': sum(r['Spare Parts'] for r in rows),
            'Fitting Charge': sum(r['Fitting Charge'] for r in rows),
            'JCB Charge': sum(r['JCB Charge'] for r in rows),
            'Cutting Wheel': sum(r['Cutting Wheel'] for r in rows),
            'Mess Expense': sum(r['Mess Expense'] for r in rows),
            'Other Expense': sum(r['Other Expense'] for r in rows),
            'Total Expense': total_expense,
            'First Quality Bricks': sum(r['First Quality Bricks'] for r in rows),
            'Second Quality Bricks': sum(r['Second Quality Bricks'] for r in rows),
            'Broken Bricks Loads': sum(r['Broken Bricks Loads'] for r in rows),
            'Total Revenue': total_revenue,
            'Land Lease Value': total_land_lease,
            'Net Value': net_value,
            'Received Amount': total_received,
            'Balance Outstanding': total_outstanding,
            'Source File': ''
        }
        rows.append(totals_row)

    df = pd.DataFrame(rows)

    if export_format == 'excel':
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Report', index=False)
            workbook = writer.book
            worksheet = writer.sheets['Report']
            
            from openpyxl.styles import Font, PatternFill, Alignment
            header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
            if len(rows) > 0:
                totals_row_idx = len(rows) + 1
                totals_font = Font(name='Arial', size=11, bold=True)
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=totals_row_idx, column=col_num)
                    cell.font = totals_font

            # Autofit column widths
            for col in worksheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = col[0].column_letter
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"{filename}.xlsx"
        )

    elif export_format == 'csv':
        csv_data = df.to_csv(index=False)
        return Response(
            csv_data,
            mimetype='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename={filename}.csv'
            }
        )

    elif export_format == 'pdf':
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=landscape(letter),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        story = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=20,
            textColor=colors.HexColor('#1F4E78'),
            spaceAfter=5
        )
        
        meta_style = ParagraphStyle(
            'MetaStyle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            textColor=colors.HexColor('#555555'),
            spaceAfter=15
        )

        story.append(Paragraph("Quarry Tracker Finance Report", title_style))
        q_text = "Quarry 1" if quarry == "Quarry 1" else ("Quarry 2" if quarry == "Quarry 2" else "Both Quarries")
        story.append(Paragraph(
            f"<b>Filter:</b> {q_text} | <b>Date Range:</b> {start_date.isoformat()} to {end_date.isoformat()} | <b>Generated:</b> {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}",
            meta_style
        ))

        # 1. Summary Block Table
        summary_data = [
            [Paragraph("<b>KPI</b>", styles['Normal']), Paragraph("<b>Value</b>", styles['Normal'])]
        ]
        kpis_list = [
            ("Total Working Days", f"{total_working_days}"),
            ("Total Revenue", f"INR {total_revenue:,.2f}"),
            ("Total Expense", f"INR {total_expense:,.2f}"),
            ("Land Lease Value", f"INR {total_land_lease:,.2f}"),
            ("Net Value", f"INR {net_value:,.2f}"),
            ("Received Amount", f"INR {total_received:,.2f}"),
            ("Balance Outstanding", f"INR {total_outstanding:,.2f}")
        ]
        for kpi, val in kpis_list:
            summary_data.append([kpi, val])
            
        t_summary = Table(summary_data, colWidths=[200, 150])
        t_summary.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F2F2F2')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#1F4E78')),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D0D0D0')),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
        ]))
        
        story.append(Paragraph("<b>Financial Summary</b>", styles['Heading3']))
        story.append(Spacer(1, 5))
        story.append(t_summary)
        story.append(Spacer(1, 15))

        # 2. Detailed Line-items Table (fit printable width 732)
        line_headers = ["Period", "Quarry", "Working Days", "Revenue", "Expense", "Lease", "Net Value", "Received", "Outstanding"]
        line_data = [line_headers]
        for p in periods:
            line_data.append([
                f"{p.day_from.isoformat()} to {p.day_to.isoformat()}",
                p.quarry,
                str(p.working_days),
                f"{p.total_revenue:,.2f}",
                f"{p.total_expense:,.2f}",
                f"{p.land_lease_value:,.2f}",
                f"{p.net_value:,.2f}",
                f"{p.received_amount:,.2f}",
                f"{p.balance_outstanding:,.2f}"
            ])
            
        if periods:
            line_data.append([
                "Total",
                "",
                str(total_working_days),
                f"{total_revenue:,.2f}",
                f"{total_expense:,.2f}",
                f"{total_land_lease:,.2f}",
                f"{net_value:,.2f}",
                f"{total_received:,.2f}",
                f"{total_outstanding:,.2f}"
            ])
            
        col_widths = [130, 60, 60, 80, 80, 80, 80, 80, 82]
        t_details = Table(line_data, colWidths=col_widths)
        t_details.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('ALIGN', (3,1), (-1,-1), 'RIGHT'),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D0D0D0')),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('BACKGROUND', (0,1), (-1,-2), colors.white) if len(periods) > 0 else ('BACKGROUND', (0,1), (-1,-1), colors.white),
            ('LINEABOVE', (0,-1), (-1,-1), 1.5, colors.HexColor('#1F4E78')),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ]))
        
        story.append(Paragraph("<b>Detailed Line Items</b>", styles['Heading3']))
        story.append(Spacer(1, 5))
        story.append(t_details)

        doc.build(story)
        pdf_buffer.seek(0)
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"{filename}.pdf"
        )

    return jsonify({'errors': ['Invalid format. Only excel, csv, pdf are supported.']}), 400
